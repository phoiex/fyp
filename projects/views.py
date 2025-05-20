from django.shortcuts import render
from django.db.models import Avg
from register.models import Project
from projects.models import Task
from projects.forms import TaskRegistrationForm
from projects.forms import ProjectRegistrationForm

# Create your views here.
def projects(request):
    projects = Project.objects.all()
    avg_projects = Project.objects.all().aggregate(Avg('complete_per'))['complete_per__avg']
    tasks = Task.objects.all()
    overdue_tasks = tasks.filter(due='2')
    context = {
        'avg_projects' : avg_projects,
        'projects' : projects,
        'tasks' : tasks,
        'overdue_tasks' : overdue_tasks,
    }
    return render(request, 'projects/projects.html', context)

def newTask(request):
    if request.method == 'POST':
        form = TaskRegistrationForm(request.POST)
        context = {'form': form}
        if form.is_valid():
            form.save()
            created = True
            context = {
                'created': created,
                'form': form,
            }
            return render(request, 'projects/new_task.html', context)
        else:
            return render(request, 'projects/new_task.html', context)
    else:
        form = TaskRegistrationForm()
        context = {
            'form': form,
        }
        return render(request,'projects/new_task.html', context)

from django.shortcuts import render, redirect
from .forms import ProjectRegistrationForm


from django.shortcuts import render, redirect
from .forms import ProjectRegistrationForm
from .models import Project  
import os
import json


def newProject(request):
    if request.method == 'POST':
        form = ProjectRegistrationForm(request.POST)
        if form.is_valid():
            project = form.save()
            project_dict = {
                'id': project.id,
                'name': project.name,
                'slug': project.slug,
                'efforts': project.efforts,
                'status': project.status,
                'dead_line': project.dead_line,
                'company_id': project.company_id,  
                'complete_per': project.complete_per,
                'description': project.description,
                'add_date': project.add_date,
                'upd_date': project.upd_date,
            }
            file_path = r'C:\Users\17905\Desktop\acdemic\UM\FYP\project-management-system-master\classproject.txt'
            with open(file_path, 'w', encoding='utf-8') as f:  
                f.write(json.dumps(project_dict, ensure_ascii=False, default=str) + '\n')

            return redirect('/projects/project-details/')
        else:
            return render(request, 'projects/new_project.html', {'form': form})
    else:
        form = ProjectRegistrationForm()
        return render(request, 'projects/new_project.html', {'form': form})


from django.shortcuts import render, get_object_or_404
from django.shortcuts import render, redirect
from .models import Project, Task
from .forms import TaskForm


def project_overview(request):
    projects = Project.objects.all().prefetch_related('task_set', 'assign')

    if request.method == 'POST':
        
        task_id = request.POST.get('task_id')
        task = Task.objects.get(id=task_id)
       
        task_form = TaskForm(request.POST, instance=task)

        if task_form.is_valid():  
            task_form.save()  
            return redirect('projects:project_overview')  

    return render(request, 'projects/project_overview.html', {'projects': projects})


from django.shortcuts import render, redirect
from .forms import ProjectDetailsForm
from .models import Project

from django.shortcuts import render, get_object_or_404
from .models import ProjectDetails, Project, Task
from .forms import ProjectDetailsForm
import json

def project_details_view(request):
    tasks = []
    selected_project = None

    if request.method == 'POST':
        form = ProjectDetailsForm(request.POST)
        if form.is_valid():
            details = form.save()
            selected_project = details.project
            tasks = Task.objects.filter(project=selected_project).prefetch_related('assign')

            # Save project data to txt
            merged_data = {
                'id': selected_project.id,
                'name': selected_project.name,
                'slug': selected_project.slug,
                'efforts': selected_project.efforts,
                'status': selected_project.status,
                'dead_line': selected_project.dead_line,
                'company_id': selected_project.company_id,
                'complete_per': selected_project.complete_per,
                'description': selected_project.description,
                'add_date': selected_project.add_date,
                'upd_date': selected_project.upd_date,
                'problem_statement': details.problem_statement,
                'project_objectives': details.project_objectives,
                'project_scope': details.project_scope,
                'goals': details.goals,
                'assumptions': details.assumptions,
                'constraints': details.constraints,
            }

            file_path = r'C:\Users\17905\Desktop\acdemic\UM\FYP\project-management-system-master\classproject.txt'
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(json.dumps(merged_data, default=str, ensure_ascii=False) + '\n')
    else:
        form = ProjectDetailsForm()
        project_id = request.GET.get('project_id')
        if project_id:
            selected_project = get_object_or_404(Project, id=project_id)
            tasks = Task.objects.filter(project=selected_project).prefetch_related('assign')

    return render(request, 'projects/project_details_form.html', {
        'form': form,
        'tasks': tasks,
        'selected_project': selected_project,
        'projects': Project.objects.all(),
    })







from django.shortcuts import render, get_object_or_404, redirect
from .models import Project, ProjectDetails, Task
from .forms import ProjectForm, TaskForm, ProjectDetailsFormedit
from django.contrib.auth.decorators import login_required

@login_required
def project_detail_edit(request):
# Get all items
    projects = Project.objects.all()
    project_data = []

    for project in projects:
        project_details = ProjectDetails.objects.get(project=project)
        tasks = Task.objects.filter(project=project)
        
        # Handle form submission
        if request.method == 'POST':
            # Create a form instance for each item
            project_form = ProjectForm(request.POST, instance=project)
            project_details_form = ProjectDetailsFormedit(request.POST, instance=project_details)
            task_forms = {task.id: TaskForm(request.POST, instance=task) for task in tasks}

            if project_form.is_valid() and project_details_form.is_valid() and all(task_form.is_valid() for task_form in task_forms.values()):
                project_form.save()
                project_details_form.save()
                for task_form in task_forms.values():
                    task_form.save()
               
        else:
            # Create a form instance when the page is first loaded
            project_form = ProjectForm(instance=project)
            project_details_form = ProjectDetailsFormedit(instance=project_details)
            task_forms = {task.id: TaskForm(instance=task) for task in tasks}

        # Saving project data and forms
        project_data.append({
            'project': project,
            'project_details': project_details,
            'tasks': tasks,
            'project_form': project_form,
            'project_details_form': project_details_form,
            'task_forms': task_forms,
        })

    context = {'project_data': project_data}
    return render(request, 'projects/project_detail.html', context)




# views.py
from django.shortcuts import render, redirect
from .forms import ProjectedInfoForm

def projected_info_add(request):
    if request.method == 'POST':
        form = ProjectedInfoForm(request.POST)
        if form.is_valid():
            # Save form data
            form.save()
            return redirect('projects:projects')  # After submitting, jump to the project list page
    else:
        form = ProjectedInfoForm()
    
    return render(request, 'projects/projected_info_add.html', {'form': form})



import subprocess
from django.shortcuts import render
from django.http import HttpResponse

# View function to handle form submission
def run_scripts(request):
    if request.method == "POST":
        try:
            # run editdatabase.py
            subprocess.run(["python", "path_to_your_script/editdatabase.py"], check=True)

            # run run-2.py
            subprocess.run(["python", "path_to_your_script/run-2.py"], check=True)

            return HttpResponse("The script runs successfully！")
        except subprocess.CalledProcessError as e:
            return HttpResponse(f"The script runs fail: {e}")

    return render(request, 'projects/project_details_form.html')


from openpyxl import Workbook
from django.shortcuts import render
from .models import Project, Task
import os

from django.shortcuts import render
from openpyxl import Workbook
import os
from .models import Project, Task, ProjectDetails, ProjectedInfo
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
def export_tasks_excel(request):
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    from openpyxl.worksheet.table import Table
    import os

    def get_first_table(ws):
        for tbl in ws._tables:
            if isinstance(tbl, Table):
                return tbl
            elif isinstance(tbl, str):
                return ws.tables.get(tbl)
        return None

    def clear_table_data_and_get_bounds(ws):
        table = get_first_table(ws)
        if not table:
            return None, None, None
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = None
        table.ref = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=min_row, column=max_col).coordinate}"
        return min_row, min_col, max_col

    projects = Project.objects.all()

    if request.method == 'POST':
        project_id = request.POST.get('project')
        project = Project.objects.get(id=project_id)
        tasks = Task.objects.filter(project=project)
        planners = Planner.objects.all()

        save_path = r"C:\Users\17905\Desktop\acdemic\UM\FYP\donload.xlsx"
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        wb = load_workbook(save_path)

        # ---- CustomerData ----
        if 'CustomerData' in wb.sheetnames:
            ws_data = wb['CustomerData']
            bounds = clear_table_data_and_get_bounds(ws_data)
            if all(bounds):
                min_row, min_col, max_col = bounds

                for row_index, task in enumerate(tasks, start=1):
                    assign_users = task.assign.all()
                    user_id = assign_users[0].id if assign_users else ''
                    assign_names = ', '.join([user.username for user in assign_users]) if assign_users else ''
                    row = [
                        user_id or 0,
                        task.id or row_index,
                        task.task_name or 0,
                        assign_names or 0,
                        task.get_status_display() if task.status else 0,
                        task.get_due_display() if task.due else 0,
                        task.task_description or 0,
                        task.start_date or 0,
                        task.due_date or 0,
                    ]
                    write_row = min_row + row_index
                    for col_idx, value in enumerate(row, start=1):
                        ws_data.cell(row=write_row, column=col_idx, value=value)

                for i in range(len(tasks) + 1, 51):
                    write_row = min_row + i
                    empty_row = [''] * 9
                    empty_row[1] = i
                    for col_idx, value in enumerate(empty_row, start=1):
                        ws_data.cell(row=write_row, column=col_idx, value=value)

                table = get_first_table(ws_data)
                if table:
                    table.ref = f"A{min_row}:I{min_row + 50}"

        # ---- CustomerInfo ----
        if 'CustomerInfo' in wb.sheetnames:
            ws_info = wb['CustomerInfo']
            bounds = clear_table_data_and_get_bounds(ws_info)
            if all(bounds):
                min_row, min_col, max_col = bounds

                for idx, planner in enumerate(planners, start=1):
                    row = [
                        idx,
                        1,
                        planner.teams_id or '',
                        planner.plannerid or '',
                        1
                    ]
                    write_row = min_row + idx
                    for col_idx, value in enumerate(row, start=1):
                        ws_info.cell(row=write_row, column=col_idx, value=value)

                table = get_first_table(ws_info)
                if table:
                    table.ref = f"A{min_row}:E{ws_info.max_row}"

        wb.save(save_path)

        return render(request, 'projects/select_project.html', {
            'projects': projects,
            'message': 'finished'
        })

    return render(request, 'projects/select_project.html', {'projects': projects})





def export_tasks_txt(request):

    projects = Project.objects.all()

    if request.method == 'POST':
        project_id = request.POST.get('project')
        project = Project.objects.get(id=project_id)


        tasks = Task.objects.filter(project=project)

        txt_save_path = r"C:\Users\17905\Desktop\acdemic\UM\FYP\project-management-system-master\update_log.txt"
        with open(txt_save_path, 'w', encoding='utf-8') as f:

            # Write basic project information
            f.write(f"Project Name: {project.name}\n")
            f.write(f"Project Description: {project.description}\n")
            f.write(f"Deadline: {project.dead_line}\n")
            f.write(f"Company: {project.company.name}\n\n")

            # Write ProjectDetails (if any)
            try:
                project_details = ProjectDetails.objects.get(project=project)
                f.write(f"Problem Statement: {project_details.problem_statement}\n")
                f.write(f"Project Objectives: {project_details.project_objectives}\n")
                f.write(f"Project Scope: {project_details.project_scope}\n")
                f.write(f"Goals: {project_details.goals}\n")
                f.write(f"Assumptions: {project_details.assumptions}\n")
                f.write(f"Constraints: {project_details.constraints}\n\n")
            except ProjectDetails.DoesNotExist:
                f.write("No Project Details available\n\n")

            try:
                projected_info = ProjectedInfo.objects.get(project=project)
                f.write(f"Business Process: {projected_info.business_process}\n")
                f.write(f"User Requirements: {projected_info.user_requirements}\n")
                f.write(f"Non-functional Requirements: {projected_info.non_functional_requirements}\n")
                f.write(f"User Roles/Permissions: {projected_info.user_roles_permissions}\n")
                f.write(f"System Integration: {projected_info.system_integration}\n")
                f.write(f"Technical Architecture: {projected_info.technical_architecture}\n")
                f.write(f"Data Management: {projected_info.data_management}\n")
                f.write(f"UI Interaction Design: {projected_info.ui_interaction_design}\n")
                f.write(f"Security Compliance: {projected_info.security_compliance}\n")
                f.write(f"Acceptance Criteria: {projected_info.acceptance_criteria}\n")
                f.write(f"Project Budget: {projected_info.project_budget}\n")
                f.write(f"Project Timeline: {projected_info.project_timeline}\n")
                f.write(f"Risk Management: {projected_info.risk_management}\n")
                f.write(f"Post Deployment Support: {projected_info.post_deployment_support}\n")
                f.write(f"Legal Compliance: {projected_info.legal_compliance}\n")
                f.write(f"Extra Information: {projected_info.extra_information}\n\n")
            except ProjectedInfo.DoesNotExist:
                f.write("No Projected Info available\n\n")


            f.write("Tasks:\n")
            for task in tasks:
                assign_users = ', '.join([user.username for user in task.assign.all()])
                f.write(f"Task Name: {task.task_name}\n")
                f.write(f"Assigned Users: {assign_users}\n")
                f.write(f"Status: {task.status}\n")
                f.write(f"Due: {task.due}\n\n")

        return render(request, 'projects/select_project.html', {'projects': projects, 'message': 'AI Report updated successfully.'})

    return render(request, 'projects/select_project.html', {'projects': projects})


from django.shortcuts import render
from .models import Project

def export_tasks(request):
    # Get all items for selection
    projects = Project.objects.all()

    # If it is a POST request, the user has selected items to export
    if request.method == 'POST':
        # Get the project ID selected by the user
        project_id = request.POST.get('project')
        project = Project.objects.get(id=project_id)
        return render(request, 'projects/select_project.html', {'projects': projects, 'message': 'Project selected for export.'})

    # If it is a GET request, display the selection interface for all items
    return render(request, 'projects/select_project.html', {'projects': projects})


from django.shortcuts import render

def dual_view(request):
    return render(request, 'projects/dual_view.html')


from django.shortcuts import render
from .forms import PlannerForm
from .models import Planner

def export_tasks_planner(request):
    success_message = ''
    if request.method == 'POST':
        form = PlannerForm(request.POST)
        if form.is_valid():
            planner_instance = Planner.objects.first()
            if planner_instance:
                planner_instance.plannerid = form.cleaned_data['plannerid']
                planner_instance.teams_id = form.cleaned_data['teams_id']
                planner_instance.save()
            else:
                form.save()
            success_message = 'submit successfully'
            form = PlannerForm() 
    else:
        form = PlannerForm()
    return render(request, 'projects/loginmicrosoft.html', {'form': form, 'success_message': success_message})


